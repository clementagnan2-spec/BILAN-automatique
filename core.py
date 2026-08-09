# -*- coding: utf-8 -*-
"""
core.py — Moteur de génération automatique du Bilan.

Principe :
1. On charge deux "balances" (Balance N-1 et Balance N), chacune une liste
   de lignes (Compte, Libellé, Débit, Crédit).
2. On calcule, pour chaque compte, son solde net = Débit - Crédit (les
   valeurs vides / non numériques comptent pour 0).
3. Le modèle du Bilan (feuille "BILAN" d'un classeur Excel) contient des
   cellules-formules utilisant des pseudo-fonctions inspirées du langage
   de l'utilisateur :

     CtaCptSoldeDébit("42*")            -> somme des comptes de racine 42
                                            dont le solde (Balance N) est
                                            débiteur (> 0)
     CtaCptSoldeCrédit("42*")           -> idem mais comptes créditeurs (<0),
                                            valeur renvoyée positive
     CtaCptSoldeDébitNm1("41*")         -> comme CtaCptSoldeDébit mais sur
                                            la Balance N-1
     CtaCptSoldeCréditNm1("41*")        -> comme CtaCptSoldeCrédit mais sur
                                            la Balance N-1
     CtaCptSolde("280*","2869*")        -> solde NET (peut être négatif),
                                            sans filtrage de sens, sur N
     CtaCptSoldeNm1(...)                -> idem sur N-1

   Avec DEUX arguments, la fonction ne fait pas l'union des deux racines
   mais la PLAGE (intervalle) entre les deux racines, bornes incluses :
   CtaCptSoldeDébit("50*","56*") = comptes dont la racine est comprise
   entre 50 et 56 inclus (50,51,...,56), pas seulement 50 et 56.

   Ces cellules-formules ne sont PAS de vraies formules Excel (Excel ne
   connaît pas ces fonctions) : ce sont de simples chaînes de texte que ce
   moteur lit, interprète et remplace par leur valeur calculée.

4. Le résultat est écrit dans une copie du classeur modèle (mêmes libellés,
   même mise en forme), avec les cellules-formules remplacées par les
   montants calculés.
"""

from __future__ import annotations

import re
import shutil
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd


# --------------------------------------------------------------------------
# 1. Chargement des balances
# --------------------------------------------------------------------------

REQUIRED_COLS = {
    "compte": ["compte", "compte général", "compte general", "n° compte", "numero compte", "numéro compte"],
    "libelle": ["libellé", "libelle", "intitulé", "intitule", "designation", "désignation"],
    "debit": ["débit", "debit", "mvt débit", "mvt debit", "solde débit", "solde debit"],
    "credit": ["crédit", "credit", "mvt crédit", "mvt credit", "solde crédit", "solde credit"],
}


def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            if v != v:  # NaN
                return 0.0
        except Exception:
            pass
        return float(v)
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return 0.0
    s = s.replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_header_row(raw: pd.DataFrame) -> int:
    """Cherche la ligne d'en-tête (celle qui contient 'compte')."""
    for i in range(min(10, len(raw))):
        row_vals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
        if any(v in REQUIRED_COLS["compte"] for v in row_vals):
            return i
    return 0


def _map_columns(header_list) -> dict:
    """Reçoit la liste des en-têtes (dans l'ordre) et renvoie {clé: index_colonne}.
    Fonctionne par POSITION plutôt que par nom, pour ne pas être piégé par des
    en-têtes dupliqués (ex. plusieurs colonnes 'Débit'/'Crédit' successives,
    typiquement report à nouveau / mouvements / solde). En cas de doublon, la
    DERNIÈRE occurrence l'emporte (c'est en général le solde final)."""
    mapping = {}
    for idx, col in enumerate(header_list):
        norm = str(col).strip().lower()
        for key, aliases in REQUIRED_COLS.items():
            if norm in aliases:
                mapping[key] = idx  # dernière occurrence gagne (écrasement volontaire)
    return mapping


@dataclass
class Balance:
    """Balance comptable agrégée : {compte(str) -> (debit, credit)}."""
    soldes: dict = field(default_factory=dict)   # compte -> solde net (debit-credit)
    libelles: dict = field(default_factory=dict)  # compte -> libellé (premier rencontré)

    def solde(self, compte: str) -> float:
        return self.soldes.get(compte, 0.0)


def load_balance(path: str, sheet_name: Optional[str] = None) -> Balance:
    """Charge un fichier de balance (xlsx/xls/csv) et renvoie un objet Balance
    avec un solde net (débit-crédit) par compte (les lignes répétées pour un
    même compte sont additionnées)."""
    path = str(path)
    if path.lower().endswith(".csv"):
        raw = pd.read_csv(path, header=None, dtype=object)
    else:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
        if isinstance(raw, dict):
            # plusieurs feuilles retournées -> prendre la première non vide
            raw = next(iter(raw.values()))

    header_row = _find_header_row(raw)
    header = raw.iloc[header_row].tolist()
    data_rows = raw.iloc[header_row + 1:]

    colmap = _map_columns(header)
    missing = [k for k in ("compte", "debit", "credit") if k not in colmap]
    if missing:
        raise ValueError(
            "Colonnes introuvables dans le fichier de balance : %s. "
            "Colonnes attendues : Compte, Libellé, Débit, Crédit." % ", ".join(missing)
        )

    bal = Balance()
    n_cols = len(header)
    for row in data_rows.itertuples(index=False, name=None):
        compte_raw = row[colmap["compte"]] if colmap["compte"] < len(row) else None
        if compte_raw is None or (isinstance(compte_raw, float) and compte_raw != compte_raw):
            continue
        compte = str(compte_raw).strip()
        if compte == "" or compte.lower() == "nan":
            continue
        # normaliser un compte du type "401100.0" -> "401100"
        if compte.endswith(".0"):
            compte = compte[:-2]

        debit = _to_float(row[colmap["debit"]]) if colmap["debit"] < len(row) else 0.0
        credit = _to_float(row[colmap["credit"]]) if colmap["credit"] < len(row) else 0.0
        lib_idx = colmap.get("libelle")
        libelle = str(row[lib_idx]) if lib_idx is not None and lib_idx < len(row) and row[lib_idx] is not None else ""

        bal.soldes[compte] = bal.soldes.get(compte, 0.0) + (debit - credit)
        if compte not in bal.libelles and libelle and libelle.lower() != "nan":
            bal.libelles[compte] = libelle

    return bal


# --------------------------------------------------------------------------
# 2. Fonctions de calcul CtaCptSolde...
# --------------------------------------------------------------------------

def _racine_range(prefixes: tuple[str, ...]) -> tuple[str, str]:
    """Normalise les préfixes (avec ou sans '*') et renvoie (debut, fin)."""
    clean = [p.rstrip("*") for p in prefixes]
    if len(clean) == 1:
        return clean[0], clean[0]
    return clean[0], clean[1]


def _compte_dans_plage(compte: str, debut: str, fin: str) -> bool:
    """Teste si `compte` a une racine comprise entre `debut` et `fin`
    (bornes incluses), plage définie sur le nombre de chiffres du préfixe
    le plus long parmi les deux bornes."""
    if debut == fin:
        return compte.startswith(debut)
    L = max(len(debut), len(fin))
    borne_min = int(debut.ljust(L, "0"))
    borne_max = int(fin.ljust(L, "9"))
    prefix_compte = compte[:L]
    if len(prefix_compte) < L:
        prefix_compte = prefix_compte.ljust(L, "0")
    try:
        val = int(prefix_compte)
    except ValueError:
        return False
    return borne_min <= val <= borne_max


def _comptes_matching(balance: Balance, prefixes: tuple[str, ...]):
    debut, fin = _racine_range(prefixes)
    for compte in balance.soldes:
        if _compte_dans_plage(compte, debut, fin):
            yield compte


def cta_cpt_solde_debit(balance: Balance, *prefixes: str) -> float:
    """Somme des soldes (positifs) des comptes débiteurs matchant les racines."""
    total = 0.0
    for compte in _comptes_matching(balance, prefixes):
        s = balance.solde(compte)
        if s > 0:
            total += s
    return total


def cta_cpt_solde_credit(balance: Balance, *prefixes: str) -> float:
    """Somme des soldes (valeur absolue) des comptes créditeurs matchant les racines."""
    total = 0.0
    for compte in _comptes_matching(balance, prefixes):
        s = balance.solde(compte)
        if s < 0:
            total += -s
    return total


def cta_cpt_solde(balance: Balance, *prefixes: str) -> float:
    """Solde net (peut être positif ou négatif), sans filtrage de sens."""
    total = 0.0
    for compte in _comptes_matching(balance, prefixes):
        total += balance.solde(compte)
    return total


# --------------------------------------------------------------------------
# 3. Évaluateur de formules
# --------------------------------------------------------------------------

# Ordre important : les noms longs (...Nm1) doivent être testés avant les
# noms courts pour ne pas être coupés par erreur lors du remplacement.
_FUNC_TOKENS = [
    ("CtaCptSoldeDébitNm1", "__F_DEBIT_NM1__"),
    ("CtaCptSoldeDebitNm1", "__F_DEBIT_NM1__"),
    ("CtaCptSoldeCréditNm1", "__F_CREDIT_NM1__"),
    ("CtaCptSoldeCreditNm1", "__F_CREDIT_NM1__"),
    ("CtaCptSoldeNm1", "__F_SOLDE_NM1__"),
    ("CtaCptSoldeDébit", "__F_DEBIT__"),
    ("CtaCptSoldeDebit", "__F_DEBIT__"),
    ("CtaCptSoldeCrédit", "__F_CREDIT__"),
    ("CtaCptSoldeCredit", "__F_CREDIT__"),
    ("CtaCptSolde", "__F_SOLDE__"),
]

# Certains modèles (liasse fiscale complète) balisent une cellule comme
# "rubrique" réutilisable ailleurs : la cellule s'écrit
#   =[R120.EtLoc]=-CtaCptSolde("10*")     (modèle Bilan)
#   =[011.EtLoc]=-CtaCptSolde("70*")      (modèles Résultat / Trésorerie...)
# ce qui signifie : "cette cellule EST la rubrique <id>, et sa valeur vaut
# <formule>". D'autres cellules peuvent ensuite réutiliser cette valeur déjà
# calculée via [<id>.EtLoc], par exemple :
#   =[R201.EtLoc]=[R120.EtLoc]+[R130.EtLoc]+...
# L'identifiant de rubrique est utilisé tel quel (avec ou sans "R"), et n'a
# pas besoin d'être unique dans tout le classeur : en cas de réutilisation
# du même identifiant sur plusieurs lignes, la dernière définition
# rencontrée (dans l'ordre des lignes) fait foi pour les références
# suivantes — ce qui correspond à l'intention des modèles observés.
_RUBRIQUE_REF_RE = re.compile(r"\[([A-Za-z0-9_]+)\.EtLoc\]")
_RUBRIQUE_ASSIGN_RE = re.compile(r"^\[([A-Za-z0-9_]+)\.EtLoc\]=(.*)$")


class FormulaError(Exception):
    pass


class RubriqueNotReady(Exception):
    """Levée quand une formule référence une rubrique [Rxxx.EtLoc] pas
    encore calculée (dépendance vers une autre cellule) : on réessaiera
    lors d'une passe ultérieure."""
    pass


def is_formula(value) -> bool:
    if not isinstance(value, str):
        return False
    v = value.strip()
    if not v.startswith("="):
        return False
    return ("CtaCptSolde" in v) or bool(_RUBRIQUE_REF_RE.search(v))


def _prepare_expr(formula: str):
    """Découpe une formule en (rubrique_id_ou_None, expression_python_prête).

    - Extrait le tag `[Rxxx.EtLoc]=` en tête de formule s'il existe.
    - Remplace les références `[Ryyy.EtLoc]` par un appel `__RUB__('Ryyy')`.
    - Remplace les noms de fonctions CtaCptSolde... par leurs tokens internes.
    """
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:]

    rubrique_id = None
    m = _RUBRIQUE_ASSIGN_RE.match(expr)
    if m:
        rubrique_id = m.group(1)
        expr = m.group(2)

    expr = _RUBRIQUE_REF_RE.sub(lambda mo: "__RUB__(%r)" % mo.group(1), expr)

    for name, token in _FUNC_TOKENS:
        expr = expr.replace(name, token)

    if "CtaCptSolde" in expr:
        raise FormulaError("Fonction non reconnue : %s" % formula)
    leftover = re.sub(r"__F_[A-Z0-9_]+__|__RUB__\([^)]*\)", "", expr)
    if re.search(r"[A-Za-zÀ-ÿ]", leftover):
        raise FormulaError("Fonction ou référence non reconnue : %s" % formula)

    return rubrique_id, expr


def _make_namespace(balance_n: "Balance", balance_n1: "Balance", rubrique_values: dict,
                     defined_ids: set = None, missing_used: set = None):
    def F_debit(*prefixes):
        return cta_cpt_solde_debit(balance_n, *prefixes)

    def F_credit(*prefixes):
        return cta_cpt_solde_credit(balance_n, *prefixes)

    def F_solde(*prefixes):
        return cta_cpt_solde(balance_n, *prefixes)

    def F_debit_nm1(*prefixes):
        return cta_cpt_solde_debit(balance_n1, *prefixes)

    def F_credit_nm1(*prefixes):
        return cta_cpt_solde_credit(balance_n1, *prefixes)

    def F_solde_nm1(*prefixes):
        return cta_cpt_solde(balance_n1, *prefixes)

    def RUB(rubrique_id):
        if rubrique_id in rubrique_values:
            return rubrique_values[rubrique_id]
        if defined_ids is not None and rubrique_id not in defined_ids:
            # Rubrique jamais définie nulle part dans la feuille (trou dans
            # le modèle, ex. ligne supprimée) : on la traite comme 0 plutôt
            # que de bloquer toute la chaîne de calcul qui en dépend.
            if missing_used is not None:
                missing_used.add(rubrique_id)
            return 0
        raise RubriqueNotReady(rubrique_id)

    return {
        "__F_DEBIT__": F_debit,
        "__F_CREDIT__": F_credit,
        "__F_SOLDE__": F_solde,
        "__F_DEBIT_NM1__": F_debit_nm1,
        "__F_CREDIT_NM1__": F_credit_nm1,
        "__F_SOLDE_NM1__": F_solde_nm1,
        "__RUB__": RUB,
    }


def evaluate_formula(formula: str, balance_n: "Balance", balance_n1: "Balance", rubrique_values: dict = None,
                      defined_ids: set = None, missing_used: set = None):
    """Évalue une formule isolée (sans dépendance de rubrique, ou avec des
    rubriques déjà connues dans `rubrique_values`). Renvoie la valeur
    numérique calculée. Lève FormulaError ou RubriqueNotReady."""
    rubrique_values = rubrique_values if rubrique_values is not None else {}
    rubrique_id, py_expr = _prepare_expr(formula)
    namespace = _make_namespace(balance_n, balance_n1, rubrique_values, defined_ids, missing_used)
    try:
        value = eval(py_expr, {"__builtins__": {}}, namespace)  # noqa: S307 (namespace restreint)
    except RubriqueNotReady:
        raise
    except Exception as e:
        raise FormulaError("Erreur d'évaluation (%s) : %s" % (formula, e))
    return value, rubrique_id


def evaluate_sheet_formulas(ws, balance_n: "Balance", balance_n1: "Balance"):
    """Évalue toutes les cellules-formules d'une feuille en plusieurs passes,
    pour résoudre les dépendances entre cellules liées par des rubriques
    [xxx.EtLoc]. Renvoie (results, errors, warnings) :
      results  = {coord: valeur}
      errors   = [(coord, formule, message)]
      warnings = [(coord, formule, message)]  (ex. rubrique jamais définie, traitée comme 0)
    """
    pending = []  # (cell, formula)
    for row in ws.iter_rows():
        for cell in row:
            if is_formula(cell.value):
                pending.append((cell, cell.value))

    # Recense à l'avance tous les identifiants de rubrique qui SERONT définis
    # quelque part dans la feuille, pour distinguer "pas encore calculé"
    # (on réessaie à la passe suivante) de "n'existe nulle part" (trou dans
    # le modèle -> valeur par défaut 0, avec un avertissement).
    defined_ids = set()
    for _cell, formula in pending:
        m = _RUBRIQUE_ASSIGN_RE.match(formula.strip()[1:] if formula.strip().startswith("=") else formula.strip())
        if m:
            defined_ids.add(m.group(1))

    rubrique_values = {}
    results = {}
    errors = []
    warnings = []
    missing_used = set()

    max_passes = len(pending) + 2
    for _ in range(max_passes):
        if not pending:
            break
        still_pending = []
        progress = False
        for cell, formula in pending:
            try:
                value, rubrique_id = evaluate_formula(
                    formula, balance_n, balance_n1, rubrique_values, defined_ids, missing_used)
            except RubriqueNotReady:
                still_pending.append((cell, formula))
                continue
            except FormulaError as e:
                errors.append((cell.coordinate, formula, str(e)))
                progress = True
                continue
            results[cell.coordinate] = value
            if rubrique_id:
                rubrique_values[rubrique_id] = value
            if missing_used:
                for rid in missing_used:
                    warnings.append((cell.coordinate, formula,
                                      "Rubrique [%s.EtLoc] jamais définie dans le modèle : traitée comme 0" % rid))
                missing_used.clear()
            progress = True
        pending = still_pending
        if not progress:
            break

    # tout ce qui reste bloqué après convergence = rubrique introuvable
    for cell, formula in pending:
        errors.append((cell.coordinate, formula, "Rubrique référencée jamais calculée (dépendance manquante)"))

    return results, errors, warnings


# --------------------------------------------------------------------------
# 4. Chargement d'un modèle Excel (xlsx natif, ou vieux format XML "SpreadsheetML")
# --------------------------------------------------------------------------

_SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"


def _is_spreadsheetml(path: str) -> bool:
    """Détecte le vieux format Excel 2003 'XML Spreadsheet' (souvent avec
    l'extension .xls alors que ce n'est pas un vrai classeur binaire)."""
    try:
        with open(path, "rb") as f:
            head = f.read(300)
        head_txt = head.decode("utf-8", errors="ignore")
        return "<?xml" in head_txt and ("mso-application" in head_txt or "Workbook" in head_txt)
    except Exception:
        return False


def _spreadsheetml_to_workbook(path: str) -> openpyxl.Workbook:
    """Convertit un fichier XML SpreadsheetML (Excel 2003) en classeur
    openpyxl équivalent (valeurs et formules-texte), sans dépendance externe."""
    tree = ET.parse(path)
    root = tree.getroot()

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    def tag(name):
        return "{%s}%s" % (_SS_NS, name)

    for ws_el in root.findall(tag("Worksheet")):
        sheet_name = ws_el.get(tag("Name")) or "Sheet"
        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        table = ws_el.find(tag("Table"))
        if table is None:
            continue

        row_idx = 0
        for row_el in table.findall(tag("Row")):
            idx_attr = row_el.get(tag("Index"))
            row_idx = int(idx_attr) if idx_attr else row_idx + 1

            col_idx = 0
            for cell_el in row_el.findall(tag("Cell")):
                idx_attr = cell_el.get(tag("Index"))
                col_idx = int(idx_attr) if idx_attr else col_idx + 1

                data_el = cell_el.find(tag("Data"))
                if data_el is not None:
                    val = data_el.text
                    dtype = data_el.get(tag("Type"), "String")
                    if val is not None and dtype == "Number":
                        try:
                            fval = float(val)
                            val = int(fval) if fval.is_integer() else fval
                        except ValueError:
                            pass
                    if val is not None:
                        ws_out.cell(row=row_idx, column=col_idx, value=val)

                span = cell_el.get(tag("MergeAcross"))
                if span:
                    col_idx += int(span)

    return wb_out


def open_template_workbook(template_path: str) -> openpyxl.Workbook:
    """Ouvre un modèle de Bilan quel que soit son format (.xlsx natif, ou
    ancien export XML 'SpreadsheetML' souvent nommé .xls)."""
    if _is_spreadsheetml(template_path):
        return _spreadsheetml_to_workbook(template_path)
    return openpyxl.load_workbook(template_path, data_only=False)


def _guess_bilan_sheet(wb: openpyxl.Workbook, preferred: str = "BILAN") -> str:
    if preferred in wb.sheetnames:
        return preferred
    # sinon, la feuille qui contient le plus de cellules-formules CtaCptSolde...
    best_name, best_count = None, -1
    for name in wb.sheetnames:
        count = 0
        for row in wb[name].iter_rows():
            for cell in row:
                if is_formula(cell.value):
                    count += 1
        if count > best_count:
            best_name, best_count = name, count
    if best_count <= 0:
        raise ValueError(
            "Aucune feuille du modèle ne contient de formules CtaCptSolde... "
            "Vérifiez le fichier modèle."
        )
    return best_name


# --------------------------------------------------------------------------
# 5. Génération des états financiers à partir d'un modèle
# --------------------------------------------------------------------------

@dataclass
class GenerationReport:
    cells_ok: int = 0
    cells_error: list = field(default_factory=list)  # [(sheet, coord, formula, message)]
    cells_warning: list = field(default_factory=list)  # [(sheet, coord, formula, message)]
    output_path: str = ""


# Registre des 4 états gérés : identifiant technique, libellé affiché, nom du
# fichier-ressource du modèle par défaut, feuille attendue dans le modèle, et
# nom de fichier de sortie suggéré.
ETATS = [
    {"id": "bilan", "label": "Bilan", "resource": "modele_bilan.xlsx",
     "sheet_hint": "BILAN", "output_name": "Bilan.xlsx"},
    {"id": "resultat", "label": "Compte de Résultat (SIG)", "resource": "modele_resultat.xlsx",
     "sheet_hint": "Feuil1", "output_name": "Compte_de_Resultat.xlsx"},
    {"id": "situation", "label": "Situation Financière (FR-BFR-TN)", "resource": "modele_situation.xlsx",
     "sheet_hint": "Feuil1", "output_name": "Situation_Financiere.xlsx"},
    {"id": "flux", "label": "Flux de Trésorerie (TFT)", "resource": "modele_flux.xlsx",
     "sheet_hint": "Feuil1", "output_name": "Flux_de_Tresorerie.xlsx"},
]


def generate_etat_from_workbook(template_path: str, balance_n: "Balance", balance_n1: "Balance",
                                 output_path: str, sheet_hint: str = "BILAN") -> GenerationReport:
    """Génère UN état financier (peu importe lequel) à partir de son modèle et
    de deux balances déjà chargées. C'est la fonction générique utilisée pour
    le Bilan comme pour le Compte de Résultat, la Situation Financière ou le
    Flux de Trésorerie — ces documents partagent le même langage de formules
    (CtaCptSolde..., rubriques [xxx.EtLoc]).
    """
    wb = open_template_workbook(template_path)
    actual_sheet = _guess_bilan_sheet(wb, preferred=sheet_hint)

    report = GenerationReport(output_path=output_path)
    ws = wb[actual_sheet]

    results, errors, warnings = evaluate_sheet_formulas(ws, balance_n, balance_n1)
    for coord, value in results.items():
        ws[coord] = value
    report.cells_ok = len(results)
    for coord, formula, msg in errors:
        report.cells_error.append((actual_sheet, coord, formula, msg))
        ws[coord] = "#ERREUR"
    for coord, formula, msg in warnings:
        report.cells_warning.append((actual_sheet, coord, formula, msg))

    wb.save(output_path)
    return report


def generate_bilan(template_path: str, balance_n1_path: str, balance_n_path: str,
                    output_path: str,
                    sheet_n1: str = None, sheet_n: str = None,
                    bilan_sheet: str = "BILAN") -> GenerationReport:
    """Génère le Bilan seul (conservé pour compatibilité). Voir
    `generate_all_etats` pour générer les 4 états d'un coup."""
    balance_n1 = load_balance(balance_n1_path, sheet_name=sheet_n1)
    balance_n = load_balance(balance_n_path, sheet_name=sheet_n)
    return generate_etat_from_workbook(template_path, balance_n, balance_n1, output_path, sheet_hint=bilan_sheet)


def generate_all_etats(balance_n1_path: str, balance_n_path: str, output_dir: str,
                        templates: dict, selected_ids: Optional[list] = None,
                        sheet_n1: str = None, sheet_n: str = None) -> dict:
    """Génère plusieurs états financiers en une seule fois, à partir des mêmes
    deux balances (chargées une seule fois pour l'efficacité).

    - templates : dict {etat_id: chemin_du_modèle} (ex. résolu par
      l'application à partir des ressources embarquées, ou remplacé par un
      modèle personnalisé choisi par l'utilisateur).
    - selected_ids : sous-ensemble d'identifiants d'états à générer parmi
      ceux du registre ETATS (par défaut : tous).
    - output_dir : dossier où écrire chaque fichier de sortie (un fichier par
      état, nommé selon `output_name` dans le registre ETATS).

    Renvoie {etat_id: GenerationReport} (ou {etat_id: exception} en cas
    d'échec total sur un état — les autres états continuent d'être générés).
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    balance_n1 = load_balance(balance_n1_path, sheet_name=sheet_n1)
    balance_n = load_balance(balance_n_path, sheet_name=sheet_n)

    selected = selected_ids if selected_ids is not None else [e["id"] for e in ETATS]
    results = {}

    for etat in ETATS:
        if etat["id"] not in selected:
            continue
        template_path = templates.get(etat["id"])
        if not template_path:
            results[etat["id"]] = FileNotFoundError(
                "Aucun modèle disponible pour '%s'." % etat["label"])
            continue
        output_path = str(Path(output_dir) / etat["output_name"])
        try:
            report = generate_etat_from_workbook(
                template_path, balance_n, balance_n1, output_path,
                sheet_hint=etat["sheet_hint"],
            )
            results[etat["id"]] = report
        except Exception as e:  # un état en échec ne doit pas bloquer les autres
            results[etat["id"]] = e

    return results


# --------------------------------------------------------------------------
# 6. Liasse Fiscale — fiche d'identification (7 champs communs à toutes les feuilles)
# --------------------------------------------------------------------------
#
# Dans le modèle de Liasse Fiscale complet, ces 7 informations ne sont PAS
# ressaisies sur chacune des ~80 feuilles : presque toutes les feuilles
# contiennent une formule qui pointe, directement ou par une chaîne de
# renvois (ex. NOTE 5!C3 -> 'NOTE 1'!C3 -> 'FICHE R4'!C3 -> GARDE!D22),
# vers une poignée de cellules "racines". Il suffit donc de renseigner ces
# cellules racines pour que l'information apparaisse partout dans la liasse.
#
# Exception : le champ "Durée (en mois)" n'existe nulle part sur la feuille
# GARDE dans ce modèle — il est saisi en dur à deux endroits indépendants
# (BILAN!N5 et FICHE R1!U5), chacun étant à l'origine de sa propre chaîne de
# renvois. Les deux doivent donc être renseignés.
#
# Autre particularité du modèle fourni : le champ "Adresse" n'est PAS
# raccordé par formule sur les feuilles FICHE R1 et FICHE R4 (cellules
# fusionnées vides). Pour que l'adresse apparaisse malgré tout partout où
# elle devrait, ces deux cellules sont également renseignées directement.

LIASSE_IDENTITY_FIELDS = [
    {"id": "denomination", "label": "Dénomination sociale de l'entité",
     "targets": [("GARDE", "D22")], "type": "text"},
    {"id": "adresse", "label": "Adresse",
     "targets": [("GARDE", "C28"), ("FICHE R1", "D4"), ("FICHE R4", "B4")], "type": "text"},
    {"id": "ifu", "label": "N° IFU du contribuable",
     "targets": [("GARDE", "D30")], "type": "text"},
    {"id": "nes", "label": "N° de télédéclarant (NES)",
     "targets": [("GARDE", "D31")], "type": "text"},
    {"id": "sigle", "label": "Sigle usuel",
     "targets": [("GARDE", "C26")], "type": "text"},
    {"id": "duree", "label": "Durée (en mois)",
     "targets": [("BILAN", "N5"), ("FICHE R1", "U5")], "type": "int"},
    {"id": "exercice_clos", "label": "Exercice clos le",
     "targets": [("GARDE", "E17")], "type": "date"},
]


def _parse_date_fr(text: str):
    text = (text or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def apply_liasse_identity(wb: openpyxl.Workbook, values: dict) -> list:
    """Écrit les 7 champs d'identité dans les cellules racines du classeur
    Liasse Fiscale (voir LIASSE_IDENTITY_FIELDS). `values` est un dict
    {field_id: texte_saisi}. Renvoie la liste des (feuille, cellule)
    réellement modifiées ; ignore silencieusement les feuilles absentes
    (modèle personnalisé incomplet) plutôt que d'échouer."""
    applied = []
    for fld in LIASSE_IDENTITY_FIELDS:
        raw = values.get(fld["id"], "")
        raw = raw.strip() if isinstance(raw, str) else raw

        value = raw
        if fld["type"] == "int" and raw not in (None, ""):
            try:
                value = int(raw)
            except (TypeError, ValueError):
                value = raw  # on laisse tel quel plutôt que planter ; l'utilisateur corrigera
        elif fld["type"] == "date" and raw not in (None, ""):
            parsed = _parse_date_fr(raw)
            value = parsed if parsed is not None else raw

        for sheet_name, coord in fld["targets"]:
            if sheet_name in wb.sheetnames:
                wb[sheet_name][coord] = value
                applied.append((sheet_name, coord))
    return applied


def generate_liasse_identity(template_path: str, values: dict, output_path: str) -> list:
    """Charge le modèle de Liasse Fiscale, y applique la fiche
    d'identification, et enregistre le résultat. Renvoie la liste des
    cellules modifiées."""
    wb = open_template_workbook(template_path)
    applied = apply_liasse_identity(wb, values)
    wb.save(output_path)
    return applied


# --------------------------------------------------------------------------
# 7. Liasse Fiscale — génération complète (BILAN + RESULTAT + TFT + identité)
# --------------------------------------------------------------------------
#
# Le modèle de Liasse Fiscale complet embarque, sur les feuilles BILAN,
# RESULTAT et TFT, les mêmes formules CtaCptSolde.../rubriques que les 4
# petits modèles — préparées à partir des comptes SYSCOHADA déjà validés.
# Cette fonction évalue ces trois feuilles (indépendamment les unes des
# autres, chacune avec son propre espace de rubriques) et applique en plus
# la fiche d'identification.

LIASSE_ETATS_SHEETS = ["BILAN", "RESULTAT", "TFT"]


def generate_liasse_complete(template_path: str, balance_n_path: str, balance_n1_path: str,
                              output_path: str, identity_values: dict = None,
                              sheet_n: str = None, sheet_n1: str = None) -> dict:
    """Génère la Liasse Fiscale complète : calcule BILAN, RESULTAT et TFT à
    partir des balances fournies, applique la fiche d'identification si
    fournie, et enregistre le résultat.

    Renvoie {sheet_name: GenerationReport} pour BILAN/RESULTAT/TFT, plus
    la clé "identite" -> liste des cellules d'identité modifiées (si
    `identity_values` est fourni).
    """
    balance_n = load_balance(balance_n_path, sheet_name=sheet_n)
    balance_n1 = load_balance(balance_n1_path, sheet_name=sheet_n1)

    wb = open_template_workbook(template_path)
    reports = {}

    for sheet_name in LIASSE_ETATS_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        report = GenerationReport(output_path=output_path)
        results, errors, warnings = evaluate_sheet_formulas(ws, balance_n, balance_n1)
        for coord, value in results.items():
            ws[coord] = value
        report.cells_ok = len(results)
        for coord, formula, msg in errors:
            report.cells_error.append((sheet_name, coord, formula, msg))
            ws[coord] = "#ERREUR"
        for coord, formula, msg in warnings:
            report.cells_warning.append((sheet_name, coord, formula, msg))
        reports[sheet_name] = report

    if identity_values:
        reports["identite"] = apply_liasse_identity(wb, identity_values)

    wb.save(output_path)
    return reports
