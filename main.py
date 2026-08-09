# -*- coding: utf-8 -*-
"""
main.py — États Financiers Automatiques
Interface graphique : importer Balance N-1 + Balance N, générer le Bilan,
le Compte de Résultat, la Situation Financière et le Flux de Trésorerie à
partir de leurs modèles Excel respectifs, en appliquant les formules
CtaCptSolde... et les rubriques [xxx.EtLoc].
"""

import io
import json
import os
import shutil
import sys
import tempfile
import traceback
import webbrowser
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import openpyxl
from openpyxl.utils import get_column_letter

import core
import security

APP_TITLE = "États Financiers Automatiques"
APP_VERSION = "2.0"


def resource_path(relative_path: str) -> str:
    """Résout le chemin d'une ressource, y compris depuis un .exe PyInstaller
    (onefile) où les fichiers sont extraits dans un dossier temporaire."""
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def user_config_dir() -> str:
    """Dossier persistant (à côté de l'exécutable) où sont enregistrés les
    modèles modifiés par l'utilisateur via le menu PARAMÈTRES. Contrairement
    au dossier `resources` embarqué dans l'exe (extrait dans un dossier
    temporaire à chaque lancement), ce dossier survit d'un lancement à
    l'autre. Les fichiers qui y sont écrits sont CHIFFRÉS (voir security.py)
    — pas de .xlsx en clair, pour qu'un utilisateur qui ouvrirait ce dossier
    dans l'Explorateur ne puisse pas lire les formules avec Excel."""
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    d = os.path.join(base, "modeles_personnalises")
    os.makedirs(d, exist_ok=True)
    return d


# Registre unifié des modèles éditables via PARAMÈTRES : les 4 états
# "simples" (core.ETATS) + la Liasse Fiscale complète (traitée à part car
# multi-feuilles). {template_id: nom_fichier_ressource}
def _template_registry() -> dict:
    reg = {etat["id"]: etat["resource"] for etat in core.ETATS}
    reg["liasse"] = LIASSE_TEMPLATE_RESOURCE
    return reg


def _custom_dat_path(template_id: str) -> str:
    registry = _template_registry()
    resource = registry.get(template_id)
    if not resource:
        return ""
    stem = os.path.splitext(resource)[0]
    return os.path.join(user_config_dir(), stem + ".dat")


def has_custom_template(template_id: str) -> bool:
    path = _custom_dat_path(template_id)
    return bool(path) and os.path.exists(path)


def load_active_workbook(template_id: str) -> openpyxl.Workbook:
    """Charge en mémoire le classeur actif pour ce modèle : la version
    personnalisée (déchiffrée) si elle existe, sinon le modèle par défaut
    embarqué."""
    dat_path = _custom_dat_path(template_id)
    if dat_path and os.path.exists(dat_path):
        with open(dat_path, "rb") as f:
            encrypted = f.read()
        decrypted = security.decrypt_bytes(encrypted)
        return openpyxl.load_workbook(io.BytesIO(decrypted), data_only=False)

    registry = _template_registry()
    resource = registry.get(template_id)
    if resource:
        default_path = resource_path(os.path.join("resources", resource))
        return core.open_template_workbook(default_path)
    raise ValueError("Modèle inconnu : %s" % template_id)


def save_custom_workbook(template_id: str, wb: openpyxl.Workbook) -> None:
    """Enregistre le classeur (chiffré) comme modèle personnalisé."""
    buf = io.BytesIO()
    wb.save(buf)
    encrypted = security.encrypt_bytes(buf.getvalue())
    dat_path = _custom_dat_path(template_id)
    os.makedirs(os.path.dirname(dat_path), exist_ok=True)
    with open(dat_path, "wb") as f:
        f.write(encrypted)


def restore_default_template(template_id: str) -> None:
    """Supprime le modèle personnalisé : revient au modèle d'origine
    embarqué (rien n'est jamais écrit en clair sur disque pour ça)."""
    dat_path = _custom_dat_path(template_id)
    if dat_path and os.path.exists(dat_path):
        os.remove(dat_path)


def materialize_temp_template(etat_id: str):
    """Écrit temporairement (en clair) le classeur actif dans un dossier
    temporaire système, pour que le moteur `core` — qui lit des fichiers sur
    disque — puisse le traiter. Renvoie (chemin_fichier, dossier_a_nettoyer) ;
    l'appelant doit supprimer `dossier_a_nettoyer` une fois la génération
    terminée (le fichier ne doit pas persister après usage)."""
    wb = load_active_workbook(etat_id)
    tmp_dir = tempfile.mkdtemp(prefix="etatsfin_")
    tmp_path = os.path.join(tmp_dir, "modele.xlsx")
    wb.save(tmp_path)
    return tmp_path, tmp_dir


# --------------------------------------------------------------------------
# Liasse Fiscale — fiche d'identification (page dédiée, sans mot de passe :
# ce ne sont pas des formules à protéger, juste des informations que
# l'utilisateur doit pouvoir modifier à tout moment)
# --------------------------------------------------------------------------

LIASSE_TEMPLATE_RESOURCE = "modele_liasse_fiscale.xlsx"


def liasse_identity_json_path() -> str:
    return os.path.join(user_config_dir(), "liasse_fiscale_identite.json")


def load_liasse_identity_values() -> dict:
    path = liasse_identity_json_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_liasse_identity_values(values: dict) -> None:
    path = liasse_identity_json_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(values, f, ensure_ascii=False, indent=2)


def materialize_temp_liasse_template():
    """Comme `materialize_temp_template`, mais pour la Liasse Fiscale
    (modèle personnalisé chiffré si présent, sinon modèle par défaut).
    Renvoie (chemin_fichier, dossier_a_nettoyer) ; l'appelant doit
    supprimer `dossier_a_nettoyer` une fois la génération terminée."""
    wb = load_active_workbook("liasse")
    tmp_dir = tempfile.mkdtemp(prefix="liassefin_")
    tmp_path = os.path.join(tmp_dir, "liasse.xlsx")
    wb.save(tmp_path)
    return tmp_path, tmp_dir


class TemplateEditorWindow(tk.Toplevel):
    """Fenêtre d'édition en grille d'un modèle d'état financier : affiche
    toutes les cellules (libellés ET formules CtaCptSolde.../rubriques,
    colonnes N comme N-1) dans une grille modifiable, avec enregistrement
    persistant."""

    def __init__(self, parent, etat_id: str, label: str = None, sheet_hint: str = None):
        super().__init__(parent)
        self.parent_app = parent
        self.etat_id = etat_id
        core_etat = next((e for e in core.ETATS if e["id"] == etat_id), None)
        if core_etat:
            self.label = core_etat["label"]
            self.sheet_hint = core_etat["sheet_hint"]
        else:
            # cas de la Liasse Fiscale (pas dans core.ETATS : plusieurs
            # feuilles pertinentes, label/feuille fournis explicitement)
            self.label = label or etat_id
            self.sheet_hint = sheet_hint or "Feuil1"

        self.title(f"PARAMÈTRES — Modèle : {self.label}")
        self.geometry("1100x650")
        self.entries = {}  # (row, col) -> tk.StringVar
        self.sheet_name = None

        self._build_ui()
        self._load_workbook()

    # ------------------------------------------------------------------
    def _build_ui(self):
        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", padx=8, pady=8)

        origin = "personnalisé (enregistré)" if has_custom_template(self.etat_id) else "par défaut (intégré)"
        ttk.Label(
            toolbar,
            text=f"État : {self.label}  —  Modèle {origin}",
            foreground="#666666",
        ).pack(side="left")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(btns, text="💾 Enregistrer", command=self._on_save).pack(side="left")
        ttk.Button(btns, text="↺ Restaurer le modèle d'origine",
                   command=self._on_restore).pack(side="left", padx=8)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left")

        ttk.Label(
            self,
            text="Modifiez librement les libellés ou les formules (ex. =CtaCptSoldeDébit(\"42*\"), "
                 "=CtaCptSoldeDébitNm1(\"42*\") pour l'année N-1, =[011.EtLoc]=... pour une rubrique). "
                 "Cliquez sur Enregistrer pour appliquer les changements.",
            wraplength=1060, justify="left", foreground="#444444",
        ).pack(anchor="w", padx=8, pady=(0, 6))

        # zone défilante contenant la grille
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.canvas = tk.Canvas(container, borderwidth=0)
        vscroll = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        hscroll = ttk.Scrollbar(container, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)

        vscroll.pack(side="right", fill="y")
        hscroll.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.grid_frame = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.grid_frame, anchor="nw")
        self.grid_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )

        self.status_var = tk.StringVar(value="")
        ttk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w").pack(fill="x", side="bottom")

    # ------------------------------------------------------------------
    def _load_workbook(self):
        try:
            wb = load_active_workbook(self.etat_id)
        except Exception as e:
            self.status_var.set(f"Erreur de chargement du modèle : {e}")
            return
        self.sheet_name = self._guess_sheet(wb)
        ws = wb[self.sheet_name]

        max_row = max(ws.max_row, 1)
        max_col = max(ws.max_column, 1)

        # en-têtes de colonnes (A, B, C...)
        ttk.Label(self.grid_frame, text="", width=5).grid(row=0, column=0)
        for c in range(1, max_col + 1):
            ttk.Label(self.grid_frame, text=get_column_letter(c), font=("Segoe UI", 9, "bold"),
                      width=28, anchor="center", relief="ridge").grid(row=0, column=c, sticky="nsew")

        for r in range(1, max_row + 1):
            ttk.Label(self.grid_frame, text=str(r), font=("Segoe UI", 9, "bold"),
                      width=5, anchor="center", relief="ridge").grid(row=r, column=0, sticky="nsew")
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                var = tk.StringVar(value="" if cell.value is None else str(cell.value))
                entry = tk.Entry(self.grid_frame, textvariable=var, width=28)
                entry.grid(row=r, column=c, sticky="nsew", padx=1, pady=1)
                self.entries[(r, c)] = var

        self.status_var.set(f"Feuille : {self.sheet_name} — {max_row} lignes × {max_col} colonnes.")

    def _guess_sheet(self, wb) -> str:
        preferred = self.sheet_hint
        if preferred in wb.sheetnames:
            return preferred
        return wb.sheetnames[0]

    # ------------------------------------------------------------------
    def _on_save(self):
        try:
            wb = load_active_workbook(self.etat_id)
            ws = wb[self.sheet_name]
            for (r, c), var in self.entries.items():
                text = var.get()
                ws.cell(row=r, column=c).value = text if text != "" else None
            save_custom_workbook(self.etat_id, wb)
        except Exception as e:
            messagebox.showerror("PARAMÈTRES", f"Échec de l'enregistrement :\n{e}")
            return

        # le modèle actif de la fenêtre principale doit refléter cette édition
        if self.etat_id in self.parent_app.template_vars:
            self.parent_app.template_vars[self.etat_id].set("")  # "" = utiliser le modèle actif (personnalisé)
        if self.etat_id in self.parent_app.template_labels:
            self.parent_app.template_labels[self.etat_id].set("Modèle personnalisé (enregistré)")

        self.status_var.set("Modèle enregistré (chiffré).")
        messagebox.showinfo("PARAMÈTRES", "Le modèle a été enregistré. "
                                           "Il sera utilisé à la prochaine génération.")

    def _on_restore(self):
        if not messagebox.askyesno(
                "PARAMÈTRES",
                "Restaurer le modèle d'origine ? Toutes vos modifications personnalisées "
                "pour cet état seront perdues."):
            return
        restore_default_template(self.etat_id)
        for widget in self.grid_frame.winfo_children():
            widget.destroy()
        self.entries.clear()
        self._load_workbook()
        if self.etat_id in self.parent_app.template_vars:
            self.parent_app.template_vars[self.etat_id].set("")
        if self.etat_id in self.parent_app.template_labels:
            self.parent_app.template_labels[self.etat_id].set("Modèle par défaut (intégré)")
        self.status_var.set("Modèle d'origine restauré.")


class LiasseIdentiteWindow(tk.Toplevel):
    """Page d'édition de la fiche d'identification de la Liasse Fiscale : les
    7 informations communes à toutes les feuilles de la liasse (dénomination,
    adresse, IFU, NES, sigle, durée, exercice clos), saisies une seule fois
    puis appliquées automatiquement partout où elles apparaissent."""

    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent

        self.title("Liasse Fiscale — Fiche d'identification")
        self.geometry("680x620")
        self.resizable(False, False)

        self.vars = {fld["id"]: tk.StringVar() for fld in core.LIASSE_IDENTITY_FIELDS}
        self.balance_n_path = tk.StringVar()
        self.balance_n1_path = tk.StringVar()

        self._build_ui()
        self._load_values()

    # ------------------------------------------------------------------
    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        ttk.Label(
            self,
            text="Ces informations apparaissent sur toutes les feuilles de la liasse "
                 "(BILAN, RESULTAT, TFT, notes annexes, fiches R1-R4...). "
                 "Renseignez-les une seule fois ici.",
            wraplength=640, justify="left",
        ).pack(anchor="w", **pad)

        form = ttk.Frame(self)
        form.pack(fill="both", expand=False, padx=12, pady=6)

        hints = {
            "exercice_clos": "format JJ/MM/AAAA, ex. 31/12/2025",
            "duree": "en mois, ex. 12",
        }

        for fld in core.LIASSE_IDENTITY_FIELDS:
            row = ttk.Frame(form)
            row.pack(fill="x", pady=6)
            ttk.Label(row, text=fld["label"] + " :", width=34, anchor="w").pack(side="left")
            entry = ttk.Entry(row, textvariable=self.vars[fld["id"]])
            entry.pack(side="left", fill="x", expand=True)
            if fld["id"] in hints:
                ttk.Label(row, text=hints[fld["id"]], foreground="#888888").pack(side="left", padx=6)

        btns_id = ttk.Frame(self)
        btns_id.pack(fill="x", padx=12, pady=(6, 12))
        ttk.Button(btns_id, text="💾 Enregistrer la fiche d'identification",
                   command=self._on_save).pack(side="left")

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=12, pady=6)

        # --- Génération complète (BILAN + RESULTAT + TFT + identité) --------
        frame_gen = ttk.LabelFrame(self, text="Générer la Liasse Fiscale complète (BILAN, RESULTAT, TFT)")
        frame_gen.pack(fill="x", padx=12, pady=6)

        ttk.Label(
            frame_gen,
            text="Calcule automatiquement le Bilan, le Compte de Résultat et le Tableau des "
                 "Flux de Trésorerie de la liasse à partir de vos balances, en plus de la "
                 "fiche d'identification ci-dessus.",
            wraplength=630, justify="left", foreground="#444444",
        ).pack(anchor="w", padx=8, pady=(6, 8))

        self._file_row(frame_gen, "Balance N :", self.balance_n_path, self._choose_balance_n)
        self._file_row(frame_gen, "Balance N-1 :", self.balance_n1_path, self._choose_balance_n1)

        ttk.Button(frame_gen, text="⚙  Générer la Liasse Fiscale complète…",
                   command=self._on_generate_complete).pack(anchor="w", padx=8, pady=(6, 10))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=12, pady=6)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left")

        self.status_var = tk.StringVar(value="")
        ttk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w",
                  wraplength=660).pack(fill="x", side="bottom")

    def _file_row(self, parent, label, var, command):
        row = ttk.Frame(parent)
        row.pack(fill="x", padx=8, pady=4)
        ttk.Label(row, text=label, width=14, anchor="w").pack(side="left")
        entry = ttk.Entry(row, textvariable=var)
        entry.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(row, text="Parcourir…", command=command).pack(side="left")

    def _choose_balance_n(self):
        path = filedialog.askopenfilename(
            title="Choisir la Balance N",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("Tous les fichiers", "*.*")],
        )
        if path:
            self.balance_n_path.set(path)

    def _choose_balance_n1(self):
        path = filedialog.askopenfilename(
            title="Choisir la Balance N-1",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("Tous les fichiers", "*.*")],
        )
        if path:
            self.balance_n1_path.set(path)

    # ------------------------------------------------------------------
    def _load_values(self):
        saved = load_liasse_identity_values()
        for fld in core.LIASSE_IDENTITY_FIELDS:
            self.vars[fld["id"]].set(saved.get(fld["id"], ""))

    def _collect_values(self) -> dict:
        return {fid: var.get() for fid, var in self.vars.items()}

    def _on_save(self):
        values = self._collect_values()
        try:
            save_liasse_identity_values(values)
        except Exception as e:
            messagebox.showerror("Liasse Fiscale", f"Échec de l'enregistrement :\n{e}")
            return
        self.status_var.set("Fiche d'identification enregistrée.")
        messagebox.showinfo("Liasse Fiscale", "Les informations ont été enregistrées.")

    def _on_generate_complete(self):
        n = self.balance_n_path.get().strip()
        n1 = self.balance_n1_path.get().strip()
        if not n or not os.path.exists(n):
            messagebox.showerror("Liasse Fiscale", "Veuillez sélectionner le fichier de la Balance N.")
            return
        if not n1 or not os.path.exists(n1):
            messagebox.showerror("Liasse Fiscale", "Veuillez sélectionner le fichier de la Balance N-1.")
            return

        values = self._collect_values()
        save_liasse_identity_values(values)

        out_path = filedialog.asksaveasfilename(
            title="Enregistrer la Liasse Fiscale complète",
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Liasse_Fiscale.xlsx",
        )
        if not out_path:
            return

        self.status_var.set("Génération en cours…")
        self.update_idletasks()

        try:
            tmp_path, tmp_dir = materialize_temp_liasse_template()
        except Exception as e:
            messagebox.showerror("Liasse Fiscale", f"Impossible de charger le modèle :\n{e}")
            self.status_var.set("Échec.")
            return

        try:
            reports = core.generate_liasse_complete(
                template_path=tmp_path,
                balance_n_path=n,
                balance_n1_path=n1,
                output_path=out_path,
                identity_values=values,
            )
        except Exception as e:
            messagebox.showerror("Liasse Fiscale", f"Échec de la génération :\n{e}")
            self.status_var.set("Échec.")
            return
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

        lines = [f"Fichier généré : {out_path}", ""]
        any_error = False
        for sheet in core.LIASSE_ETATS_SHEETS:
            r = reports.get(sheet)
            if r is None:
                continue
            lines.append(f"— {sheet} — {r.cells_ok} cellules calculées"
                         + (f", {len(r.cells_error)} erreurs" if r.cells_error else ", aucune erreur"))
            if r.cells_error:
                any_error = True
        if "identite" in reports:
            lines.append(f"— Fiche d'identification — {len(reports['identite'])} cellules mises à jour")

        self.status_var.set(" | ".join(lines[1:]))
        messagebox.showinfo(
            "Liasse Fiscale",
            "\n".join(lines) + (
                "\n\nDes erreurs sont survenues sur certaines cellules (voir le fichier, "
                "marquées #ERREUR)." if any_error else
                "\n\nRappel : le mapping comptable de ces 3 états est un premier jet basé sur "
                "les comptes SYSCOHADA standards, à faire vérifier avant tout usage officiel."
            ),
        )


class EtatsApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} — v{APP_VERSION}")
        self.geometry("780x640")
        self.minsize(700, 560)

        self.balance_n1_path = tk.StringVar()
        self.balance_n_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.status_var = tk.StringVar(value="Prêt.")

        # une case à cocher + un état par état
        self.etat_vars = {}         # id -> BooleanVar (généré ou non)
        self.template_vars = {}     # id -> StringVar ("" = modèle actif interne, sinon chemin externe choisi)
        self.template_labels = {}   # id -> StringVar (texte affiché : "par défaut" / "personnalisé" / chemin externe)

        self._unlocked_role = None  # None | "utilisateur" | "admin" — se reverrouille à chaque relance

        self._build_menu()
        self._build_ui()

    # ------------------------------------------------------------------
    def _build_menu(self):
        menubar = tk.Menu(self)

        liasse_menu = tk.Menu(menubar, tearoff=0)
        liasse_menu.add_command(label="Fiche d'identification…", command=self._open_liasse_identite)
        menubar.add_cascade(label="LIASSE FISCALE", menu=liasse_menu)

        parametres_menu = tk.Menu(menubar, tearoff=0)
        for etat in core.ETATS:
            parametres_menu.add_command(
                label=etat["label"],
                command=lambda eid=etat["id"]: self._open_template_editor(eid),
            )

        parametres_menu.add_separator()
        liasse_params_menu = tk.Menu(parametres_menu, tearoff=0)
        for sheet_name in core.LIASSE_ETATS_SHEETS:
            liasse_params_menu.add_command(
                label=sheet_name,
                command=lambda s=sheet_name: self._open_liasse_template_editor(s),
            )
        parametres_menu.add_cascade(label="Liasse Fiscale (BILAN / RESULTAT / TFT)", menu=liasse_params_menu)

        parametres_menu.add_separator()
        parametres_menu.add_command(label="🔑 Mot de passe utilisateur du mois (Admin)",
                                     command=self._show_user_password)
        parametres_menu.add_command(label="Verrouiller", command=self._lock_parametres)
        menubar.add_cascade(label="PARAMÈTRES", menu=parametres_menu)

        self.config(menu=menubar)

    def _open_liasse_identite(self):
        LiasseIdentiteWindow(self)

    def _open_liasse_template_editor(self, sheet_name: str):
        if not self._ensure_unlocked():
            return
        TemplateEditorWindow(self, "liasse", label=f"Liasse Fiscale — {sheet_name}", sheet_hint=sheet_name)

    def _lock_parametres(self):
        self._unlocked_role = None
        messagebox.showinfo("PARAMÈTRES", "Accès reverrouillé.")

    def _ensure_unlocked(self) -> bool:
        """Demande un mot de passe (utilisateur du mois OU administrateur) si
        l'accès à PARAMÈTRES n'a pas déjà été déverrouillé dans cette
        session. Renvoie True si l'accès est autorisé."""
        if self._unlocked_role is not None:
            return True
        pwd = simpledialog.askstring(
            "PARAMÈTRES — Accès protégé",
            "Mot de passe :",
            show="*",
            parent=self,
        )
        if pwd is None:
            return False  # annulé
        ok, role = security.check_any_password(pwd)
        if ok:
            self._unlocked_role = role
            return True
        messagebox.showerror("PARAMÈTRES", "Mot de passe incorrect.")
        return False

    def _ensure_admin(self) -> bool:
        """Exige spécifiquement le mot de passe Administrateur (même si un
        accès « utilisateur » est déjà déverrouillé dans la session)."""
        if self._unlocked_role == "admin":
            return True
        pwd = simpledialog.askstring(
            "Accès Administrateur",
            "Mot de passe administrateur :",
            show="*",
            parent=self,
        )
        if pwd is None:
            return False
        if security.check_admin_password(pwd):
            self._unlocked_role = "admin"
            return True
        messagebox.showerror("PARAMÈTRES", "Mot de passe administrateur incorrect.")
        return False

    def _show_user_password(self):
        if not self._ensure_admin():
            return
        now = datetime.now()
        lines = []
        for i in range(3):
            year = now.year + (now.month - 1 + i) // 12
            month = (now.month - 1 + i) % 12 + 1
            when = datetime(year, month, 1)
            pwd = security.generate_monthly_password(when=when)
            suffix = "  ← ce mois-ci" if i == 0 else ""
            lines.append(f"{when.strftime('%Y-%m')} : {pwd}{suffix}")
        messagebox.showinfo(
            "Mot de passe utilisateur du mois",
            "Communiquez uniquement le mot de passe du mois en cours à vos utilisateurs :\n\n"
            + "\n".join(lines),
        )

    def _open_template_editor(self, etat_id: str):
        if not self._ensure_unlocked():
            return
        TemplateEditorWindow(self, etat_id)

    # ------------------------------------------------------------------

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        header = ttk.Frame(self)
        header.pack(fill="x", **pad)
        ttk.Label(header, text=APP_TITLE, font=("Segoe UI", 16, "bold")).pack(anchor="w")
        ttk.Label(
            header,
            text="Importez la Balance N-1 et la Balance N (colonnes Compte, Libellé, "
                 "Débit, Crédit) pour générer automatiquement les états financiers "
                 "sélectionnés ci-dessous.",
            wraplength=740, justify="left",
        ).pack(anchor="w", pady=(4, 0))

        # --- Balances --------------------------------------------------
        frame_n1 = ttk.LabelFrame(self, text="Balance N-1 (exercice précédent)")
        frame_n1.pack(fill="x", **pad)
        self._file_row(frame_n1, self.balance_n1_path, self._choose_balance_n1)

        frame_n = ttk.LabelFrame(self, text="Balance N (exercice en cours)")
        frame_n.pack(fill="x", **pad)
        self._file_row(frame_n, self.balance_n_path, self._choose_balance_n)

        # --- États à générer --------------------------------------------
        frame_etats = ttk.LabelFrame(self, text="États à générer")
        frame_etats.pack(fill="x", **pad)
        for etat in core.ETATS:
            row = ttk.Frame(frame_etats)
            row.pack(fill="x", padx=8, pady=4)

            var = tk.BooleanVar(value=True)
            self.etat_vars[etat["id"]] = var
            ttk.Checkbutton(row, text=etat["label"], variable=var, width=32).pack(side="left")

            self.template_vars[etat["id"]] = tk.StringVar(value="")  # "" = modèle actif interne
            label_text = ("Modèle personnalisé (enregistré)" if has_custom_template(etat["id"])
                          else "Modèle par défaut (intégré)")
            lvar = tk.StringVar(value=label_text)
            self.template_labels[etat["id"]] = lvar
            ttk.Label(row, textvariable=lvar, foreground="#666666", width=32).pack(
                side="left", fill="x", expand=True, padx=6)
            ttk.Button(row, text="Modèle externe…",
                       command=lambda eid=etat["id"]: self._choose_template(eid)).pack(side="left")

        ttk.Label(
            frame_etats,
            text="Le modèle par défaut (ou celui édité via PARAMÈTRES) est utilisé automatiquement. "
                 "« Modèle externe… » permet de charger ponctuellement un autre fichier (accès protégé).",
            foreground="#666666",
        ).pack(anchor="w", padx=8, pady=(0, 6))

        # --- Sortie -----------------------------------------------------
        frame_out = ttk.LabelFrame(self, text="Dossier de sortie")
        frame_out.pack(fill="x", **pad)
        self._file_row(frame_out, self.output_dir, self._choose_output_dir)

        # --- Bouton générer ----------------------------------------------
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", **pad)
        self.generate_btn = ttk.Button(btn_frame, text="⚙  Générer les états sélectionnés",
                                        command=self._on_generate)
        self.generate_btn.pack(side="left")
        ttk.Button(btn_frame, text="Ouvrir le dossier de sortie", command=self._open_output_dir).pack(
            side="left", padx=8)

        # --- Journal / résultats ------------------------------------------
        frame_log = ttk.LabelFrame(self, text="Résultat")
        frame_log.pack(fill="both", expand=True, **pad)
        self.log_text = tk.Text(frame_log, height=14, wrap="word", state="disabled")
        self.log_text.pack(fill="both", expand=True, padx=6, pady=6)

        status_bar = ttk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w")
        status_bar.pack(fill="x", side="bottom")

    def _file_row(self, parent, var, command):
        row = ttk.Frame(parent)
        row.pack(fill="x", padx=8, pady=6)
        entry = ttk.Entry(row, textvariable=var)
        entry.pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Parcourir…", command=command).pack(side="left", padx=6)

    # ------------------------------------------------------------------
    def _choose_balance_n1(self):
        path = filedialog.askopenfilename(
            title="Choisir la Balance N-1",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("Tous les fichiers", "*.*")],
        )
        if path:
            self.balance_n1_path.set(path)
            if not self.output_dir.get():
                self.output_dir.set(os.path.join(os.path.dirname(path), "Etats_financiers"))

    def _choose_balance_n(self):
        path = filedialog.askopenfilename(
            title="Choisir la Balance N",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("Tous les fichiers", "*.*")],
        )
        if path:
            self.balance_n_path.set(path)
            if not self.output_dir.get():
                self.output_dir.set(os.path.join(os.path.dirname(path), "Etats_financiers"))

    def _choose_template(self, etat_id: str):
        if not self._ensure_unlocked():
            return
        path = filedialog.askopenfilename(
            title="Choisir le modèle",
            filetypes=[("Classeurs Excel", "*.xlsx *.xls"), ("Tous les fichiers", "*.*")],
        )
        if path:
            self.template_vars[etat_id].set(path)
            self.template_labels[etat_id].set(f"Modèle externe : {os.path.basename(path)}")

    def _choose_output_dir(self):
        path = filedialog.askdirectory(title="Choisir le dossier de sortie")
        if path:
            self.output_dir.set(path)

    # ------------------------------------------------------------------
    def _log(self, text: str, clear: bool = False):
        self.log_text.configure(state="normal")
        if clear:
            self.log_text.delete("1.0", "end")
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _on_generate(self):
        n1 = self.balance_n1_path.get().strip()
        n = self.balance_n_path.get().strip()
        out_dir = self.output_dir.get().strip()

        if not n1 or not os.path.exists(n1):
            messagebox.showerror(APP_TITLE, "Veuillez sélectionner le fichier de la Balance N-1.")
            return
        if not n or not os.path.exists(n):
            messagebox.showerror(APP_TITLE, "Veuillez sélectionner le fichier de la Balance N.")
            return
        if not out_dir:
            messagebox.showerror(APP_TITLE, "Veuillez indiquer un dossier de sortie.")
            return

        selected_ids = [eid for eid, var in self.etat_vars.items() if var.get()]
        if not selected_ids:
            messagebox.showerror(APP_TITLE, "Sélectionnez au moins un état à générer.")
            return

        # Prépare les modèles : un fichier externe explicitement choisi est
        # utilisé tel quel ; sinon, le modèle actif (personnalisé chiffré, ou
        # par défaut) est déchiffré dans un fichier temporaire qui sera
        # supprimé juste après la génération.
        templates = {}
        temp_dirs_to_clean = []
        for eid in selected_ids:
            external = self.template_vars[eid].get().strip()
            if external:
                if not os.path.exists(external):
                    messagebox.showerror(APP_TITLE, f"Modèle externe introuvable pour « {eid} ».")
                    return
                templates[eid] = external
            else:
                try:
                    tmp_path, tmp_dir = materialize_temp_template(eid)
                except Exception as e:
                    messagebox.showerror(APP_TITLE, f"Impossible de charger le modèle pour « {eid} » :\n{e}")
                    for d in temp_dirs_to_clean:
                        shutil.rmtree(d, ignore_errors=True)
                    return
                templates[eid] = tmp_path
                temp_dirs_to_clean.append(tmp_dir)

        self.status_var.set("Génération en cours…")
        self.generate_btn.configure(state="disabled")
        self.update_idletasks()

        try:
            try:
                results = core.generate_all_etats(
                    balance_n1_path=n1,
                    balance_n_path=n,
                    output_dir=out_dir,
                    templates=templates,
                    selected_ids=selected_ids,
                )
            except Exception as e:
                self._log("ERREUR : " + str(e), clear=True)
                self._log(traceback.format_exc())
                messagebox.showerror(APP_TITLE, f"Échec de la génération :\n{e}")
                self.status_var.set("Échec.")
                self.generate_btn.configure(state="normal")
                return
        finally:
            # les fichiers-modèles déchiffrés en clair ne doivent jamais persister
            for d in temp_dirs_to_clean:
                shutil.rmtree(d, ignore_errors=True)

        self._log(f"Dossier de sortie : {out_dir}", clear=True)
        self._log("")
        any_error = False
        for etat in core.ETATS:
            if etat["id"] not in results:
                continue
            r = results[etat["id"]]
            self._log(f"— {etat['label']} —")
            if isinstance(r, Exception):
                any_error = True
                self._log(f"  ÉCHEC : {r}")
            else:
                self._log(f"  Fichier : {r.output_path}")
                self._log(f"  Cellules calculées : {r.cells_ok}")
                if r.cells_warning:
                    self._log(f"  Avertissements : {len(r.cells_warning)}")
                    for sheet, coord, formula, msg in r.cells_warning:
                        self._log(f"    - [{sheet}!{coord}] {msg}")
                if r.cells_error:
                    any_error = True
                    self._log(f"  Erreurs : {len(r.cells_error)}")
                    for sheet, coord, formula, msg in r.cells_error:
                        self._log(f"    - [{sheet}!{coord}] {formula}\n        -> {msg}")
                else:
                    self._log("  Aucune erreur.")
            self._log("")

        self.status_var.set("Terminé.")
        self.generate_btn.configure(state="normal")
        messagebox.showinfo(
            APP_TITLE,
            "Génération terminée." if not any_error else "Génération terminée, avec des erreurs sur certains états (voir le détail)."
        )

    def _open_output_dir(self):
        out_dir = self.output_dir.get().strip()
        if not out_dir or not os.path.isdir(out_dir):
            messagebox.showwarning(APP_TITLE, "Aucun dossier de sortie valide pour l'instant.")
            return
        try:
            os.startfile(out_dir)  # type: ignore[attr-defined]
        except AttributeError:
            webbrowser.open(f"file://{out_dir}")


def main():
    app = EtatsApp()
    app.mainloop()


if __name__ == "__main__":
    main()
